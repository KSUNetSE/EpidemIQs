import os, re, json, csv
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook
from sentence_transformers import SentenceTransformer
import numpy as np
import argparse
from typing import Any, Dict, List, Optional, Tuple
import numpy as np
from openpyxl import Workbook, load_workbook
from sentence_transformers import SentenceTransformer



BIB_ENV_RE = re.compile(
    r"(\\begin\{thebibliography\}\{[^}]*\})(.*?)(\\end\{thebibliography\})",
    flags=re.DOTALL
)

BIBITEM_RE = re.compile(
    r"(\\bibitem\{(?P<key>[^}]+)\})(?P<body>.*?)(?=(\\bibitem\{)|\\end\{thebibliography\})",
    flags=re.DOTALL
)

LATEX_CMD_RE = re.compile(r"\\[a-zA-Z]+\*?(?:\[[^\]]*\])?(?:\{[^{}]*\})?")

def strip_latex(s: str) -> str:
    # common: \textit{...} -> ...
    s = re.sub(r"\\textit\{([^}]*)\}", r"\1", s)
    s = re.sub(r"\\emph\{([^}]*)\}", r"\1", s)
    # remove other commands
    s = re.sub(LATEX_CMD_RE, " ", s)
    # remove braces
    s = s.replace("{", " ").replace("}", " ")
    # normalize whitespace
    s = re.sub(r"\s+", " ", s).strip()
    return s

def parse_bibitems(tex: str) -> List[Dict[str, str]]:
    m = BIB_ENV_RE.search(tex)
    if not m:
        raise ValueError("No \\begin{thebibliography}{...} ... \\end{thebibliography} found in the .tex file.")

    body = m.group(2)
    items = []
    for im in BIBITEM_RE.finditer(body):
        key = im.group("key").strip()
        raw_body = im.group("body").strip()
        items.append({
            "bibkey": key,
            "raw": raw_body,
            "reference": strip_latex(raw_body)
        })
    return items

def flatten_to_ref_texts(obj: Any) -> List[str]:
    """
    Robustly extract reference-like strings from many JSON schemas, including:
    - { "results": [ { "papers": [ { "title": ..., "citation": ... }, ... ] }, ... ] }
    - list of dict records
    - dict of dict records
    - dict with "references": [...]
    """
    texts: List[str] = []

    def add_record(d: Dict[str, Any]) -> None:
        # Prefer full citation if present (your JSON has this)
        citation = d.get("citation") or d.get("reference") or d.get("citation_text") or ""
        title = d.get("title") or d.get("paper_title") or d.get("name") or ""
        authors = d.get("authors") or d.get("author") or d.get("creators") or ""
        venue = d.get("venue") or d.get("journal") or d.get("booktitle") or d.get("publisher") or ""
        year = d.get("year") or d.get("date") or ""

        if isinstance(authors, list):
            authors = ", ".join(str(a) for a in authors)

        # Build one semantic string per record
        parts = [str(citation).strip(), str(title).strip(), str(authors).strip(),
                 str(venue).strip(), str(year).strip()]
        text = strip_latex(" ".join([p for p in parts if p]).strip())

        if text:
            texts.append(text)

    def walk(x: Any) -> None:
        if isinstance(x, dict):
            # handle your schema explicitly
            if "papers" in x and isinstance(x["papers"], list):
                for p in x["papers"]:
                    walk(p)
            if "results" in x and isinstance(x["results"], list):
                for r in x["results"]:
                    walk(r)

            # if this dict looks like a paper/reference record, collect it
            if any(k in x for k in ("citation", "reference", "title", "paper_title", "bibtex")):
                add_record(x)

            # continue recursion through remaining fields
            for v in x.values():
                walk(v)

        elif isinstance(x, list):
            for it in x:
                walk(it)

        # ignore primitives

    walk(obj)

    # de-duplicate while preserving order
    seen = set()
    uniq = []
    for t in texts:
        if t not in seen:
            seen.add(t)
            uniq.append(t)
    return uniq

def append_to_excel(excel_path: str, tex_path: str, fraction_not_found: float) -> None:
    """
    Excel format requested:
    first row (header): tex_path, fraction_not_found
    then one row per run.
    """
    header = ["tex_path", "fraction_not_found"]

    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
        # ensure header exists
        if ws.max_row < 1 or ws.cell(1, 1).value != "tex_path" or ws.cell(1, 2).value != "fraction_not_found":
            ws.insert_rows(1)
            ws.cell(1, 1).value = header[0]
            ws.cell(1, 2).value = header[1]
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "summary"
        ws.append(header)

    ws.append([tex_path, float(fraction_not_found)])
    wb.save(excel_path)

def semantic_similarity_percent(query_text: str, lib_emb: np.ndarray, model: SentenceTransformer) -> float:
    """
    Returns best cosine similarity as a percentage [0..100].
    lib_emb: normalized embeddings (N, d)
    """
    q_emb = model.encode([query_text], normalize_embeddings=True)
    sims = lib_emb @ q_emb[0]  # cosine similarities (N,)
    best = float(np.max(sims)) if sims.size else 0.0
    best = max(0.0, best)  # clamp negatives
    return 100.0 * best

def return_not_found_references(
    references_text: str,
    json_paths: list[str],
    *,
    threshold_percent: float = 85.0,
    model_name: str = "all-MiniLM-L6-v2",
) -> list[str]:
    """
    Input:
      - references_text: str containing ONLY references (can include \\bibitem blocks or plain refs)
      - json_paths: list of paths to JSON literature files (some may not exist)
    Output:
      - list[str] of references considered NOT FOUND in the JSON library (semantic similarity < threshold_percent)

    Notes:
      - Missing JSON files are ignored.
      - If no usable refs can be extracted from existing JSON files, returns ALL input references as not found.
      - Requires: pip install sentence-transformers numpy
    """
    import os, re, json
    import numpy as np
    from sentence_transformers import SentenceTransformer

    # ---------- helpers ----------
    BIBITEM_RE = re.compile(
        r"(\\bibitem\{(?P<key>[^}]+)\})(?P<body>.*?)(?=(\\bibitem\{)|\Z)",
        flags=re.DOTALL
    )
    LATEX_CMD_RE = re.compile(r"\\[a-zA-Z]+\*?(?:\[[^\]]*\])?(?:\{[^{}]*\})?")
    YEAR_RE = re.compile(r"\b(19|20)\d{2}\b")

    def strip_latex(s: str) -> str:
        s = re.sub(r"\\textit\{([^}]*)\}", r"\1", s)
        s = re.sub(r"\\emph\{([^}]*)\}", r"\1", s)
        s = re.sub(LATEX_CMD_RE, " ", s)
        s = s.replace("{", " ").replace("}", " ")
        s = re.sub(r"\s+", " ", s).strip()
        return s

    def parse_references_block(s: str) -> list[str]:
        s = s.strip()
        if not s:
            return []
        # If it contains \bibitem, extract bodies
        if "\\bibitem" in s:
            out = []
            for m in BIBITEM_RE.finditer(s):
                body = m.group("body").strip()
                if body:
                    out.append(strip_latex(body))
            return [x for x in out if x]
        # Otherwise: split by blank lines (common for plain reference lists)
        chunks = [strip_latex(x) for x in re.split(r"\n\s*\n+", s) if x.strip()]
        # fallback: one-per-line if still too few
        if len(chunks) <= 1:
            chunks = [strip_latex(x) for x in s.splitlines() if x.strip()]
        return [x for x in chunks if x]

    def flatten_json_to_ref_texts(obj: object) -> list[str]:
        texts: list[str] = []

        def add_text(val: str) -> None:
            t = strip_latex(str(val)).strip()
            if t:
                texts.append(t)

        def add_record(d: dict) -> None:
            # Prefer full citation fields
            for k in ("citation", "reference", "citation_text", "bibtex"):
                v = d.get(k)
                if isinstance(v, str) and v.strip():
                    add_text(v)
                    return

            title = d.get("title") or d.get("paper_title") or d.get("name") or ""
            authors = d.get("authors") or d.get("author") or d.get("creators") or ""
            venue = d.get("venue") or d.get("journal") or d.get("booktitle") or d.get("publisher") or ""
            year = d.get("year") or d.get("date") or ""

            if isinstance(authors, list):
                authors = ", ".join(str(a) for a in authors)

            combo = " ".join([str(x).strip() for x in (title, authors, venue, year) if str(x).strip()])
            if combo:
                add_text(combo)

        def walk(x: object) -> None:
            if isinstance(x, dict):
                # common nestings
                for k in ("results", "papers", "references", "items", "data"):
                    v = x.get(k)
                    if isinstance(v, list):
                        for it in v:
                            walk(it)
                    elif isinstance(v, dict):
                        walk(v)

                if any(k in x for k in ("citation", "reference", "title", "paper_title", "bibtex")):
                    add_record(x)

                for v in x.values():
                    walk(v)

            elif isinstance(x, list):
                for it in x:
                    walk(it)

            elif isinstance(x, str):
                add_text(x)

        walk(obj)

        # de-dup preserving order
        seen = set()
        uniq = []
        for t in texts:
            if t not in seen:
                seen.add(t)
                uniq.append(t)

        # keep only "reference-like" strings (reduce noise)
        ref_like = [t for t in uniq if YEAR_RE.search(t) and len(t) >= 40]
        return ref_like

    # ---------- load JSON library ----------
    existing_paths = [p for p in json_paths if isinstance(p, str) and os.path.exists(p)]
    if not existing_paths:
        # no library to match against
        return parse_references_block(references_text)

    library_texts: list[str] = []
    for p in existing_paths:
        with open(p, "r", encoding="utf-8") as f:
            obj = json.load(f)
        library_texts.extend(flatten_json_to_ref_texts(obj))

    # If nothing usable extracted, treat all as not found
    refs = parse_references_block(references_text)
    if not library_texts:
        return refs

    # ---------- semantic match ----------
    model = SentenceTransformer(model_name)
    lib_emb = model.encode(library_texts, normalize_embeddings=True)

    not_found: list[str] = []
    for r in refs:
        q = strip_latex(r)
        q_emb = model.encode([q], normalize_embeddings=True)[0]
        sims = lib_emb @ q_emb
        best = float(np.max(sims)) if sims.size else 0.0
        best_pct = 100.0 * max(0.0, best)
        if best_pct < threshold_percent:
            not_found.append(r)

    return not_found


def run_ref_check(
    tex_path: str,
    json1_path: str,
    json2_path: str,
    threshold: float = 70.0,
    out_json: Optional[str] = None,
    out_excel: str = "hallucination_summary.xlsx",
    model_name: str = "all-MiniLM-L6-v2",
) -> Tuple[str, str]:
    tex = open(tex_path, "r", encoding="utf-8").read()
    bibitems = parse_bibitems(tex)

    lib_texts: List[str] = []
    for p in [json1_path, json2_path]:
        obj = json.load(open(p, "r", encoding="utf-8"))
        lib_texts.extend(flatten_to_ref_texts(obj))

    if not lib_texts:
        raise ValueError("No usable reference texts extracted from the JSON files. (Unexpected JSON schema.)")

    model = SentenceTransformer(model_name)
    lib_emb = model.encode(lib_texts, normalize_embeddings=True)

    hallucinated_count = 0
    items_out = []

    for b in bibitems:
        ref_text = b["reference"]
        sim_pct = semantic_similarity_percent(ref_text, lib_emb, model)
        hallucinated = bool(sim_pct < threshold)
        if hallucinated:
            hallucinated_count += 1

        items_out.append({
            "reference": ref_text,
            "not-among-retrieved-papers": hallucinated
        })

    total_refs = len(bibitems)
    fraction_not_found = (hallucinated_count / total_refs) if total_refs else 0.0

    if out_json is None:
        base = os.path.splitext(os.path.basename(tex_path))[0]
        out_json = f"{base}.refcheck.json"

    payload = {
        "tex_path": tex_path,
        "total_refs": total_refs,
        "hallucinated_refs": hallucinated_count,
        "fraction_not_found": fraction_not_found,
        "threshold_percent": float(threshold),
        "items": items_out
    }

    with open(out_json, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    append_to_excel(out_excel, tex_path, fraction_not_found)

    return out_json, out_excel


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("tex_path")
    ap.add_argument("json1_path")
    ap.add_argument("json2_path")
    ap.add_argument("--threshold", type=float, default=70.0)
    ap.add_argument("--out-json", default=None)
    ap.add_argument("--out-excel", default="hallucination_summary.xlsx")
    ap.add_argument("--model", default="all-MiniLM-L6-v2")
    args = ap.parse_args()

    out_json, out_excel = run_ref_check(
        tex_path=args.tex_path,
        json1_path=args.json1_path,
        json2_path=args.json2_path,
        threshold=args.threshold,
        out_json=args.out_json,
        out_excel=args.out_excel,
        model_name=args.model,
    )

    print("Done.")
    print(f"JSON saved:   {out_json}")
    print(f"Excel saved:  {out_excel}")


if __name__ == "__main__":
    from pathlib import Path
    """
    def list_folders(directory: str = ".") -> list[str]:
        p = Path(directory).expanduser().resolve()
        if not p.exists():
            raise FileNotFoundError(f"Directory not found: {p}")
        if not p.is_dir():
            raise NotADirectoryError(f"Not a directory: {p}")

        folders = sorted([x.name for x in p.iterdir() if x.is_dir()])
        return folders
    root_dir = "user-files/Q1"
    print(list_folders(root_dir))
    root_dir = "user-files/Q1"
    for folder in list_folders(root_dir):
        print(f"Processing folder: {folder}")
        try:
            path_tex = os.path.join(root_dir, folder, "initial_report.tex")
            run_ref_check(tex_path=path_tex,
                json1_path=os.path.join(root_dir, folder, "literature_review.json"),
                json2_path=os.path.join(root_dir, folder, "back_grounds.json"),
                threshold=75.0,
            out_json=None,
                out_excel="hallucination_summary.xlsx",
                model_name="all-MiniLM-L6-v2")
        except Exception as e:  
            print(f"Error processing folder {folder}: {e}")"""
    references=r"""
\begin{thebibliography}{99}

\bibitem{PastorSatorrasVespignani2001} R. Pastor-Satorras and Alessandro Vespignani. Immunization of complex networks. Physical Review E, Statistical, Nonlinear, and Soft Matter Physics, 2001.

\bibitem{CohenHavlinBenAvraham2002} R. Cohen, S. Havlin, and D. ben-Avraham. Efficient immunization strategies for computer networks and populations. Physical Review Letters, 2002.

\bibitem{MoroneMakse2015} F. Morone and H. Makse. Influence maximization in complex networks through optimal percolation. Nature, 2015.

\bibitem{FuSmallWalker2008} Xinchu Fu, M. Small, D. Walker, et al. Epidemic dynamics on scale-free networks with piecewise linear infectivity and immunization. Physical Review E, Statistical, Nonlinear, and Soft Matter Physics, 2008.

\bibitem{ParshaniCarmiHavlin2010} Roni Parshani, S. Carmi, and S. Havlin. Epidemic threshold for the susceptible-infectious-susceptible model on random networks. Physical Review Letters, 2010.

\bibitem{LeeShimNoh2012} Hyun Keun Lee, Pyoung-seop Shim, and J. Noh. Epidemic threshold of the susceptible-infected-susceptible model on complex networks. Physical Review E, Statistical, Nonlinear, and Soft Matter Physics, 2012.

\bibitem{WangWangLiu2015} Wei Wang, Wen Wang, Quan-Hui Liu, et al. Predicting the epidemic threshold of the susceptible-infected-recovered model. Scientific Reports, 2015.

\bibitem{FerreyraJonckheerePinasco2019} Emanuel Javier Ferreyra, M. Jonckheere, J. P. Pinasco. SIR Dynamics with Vaccination in a Large Configuration Model. Applied Mathematics and Optimization, 2019.

\bibitem{MatsukiTanaka2019} Akari Matsuki, G. Tanaka. Intervention threshold for epidemic control in susceptible-infected-recovered metapopulation models. Physical Review E, 2019.

\bibitem{Dadashkarimi2025} M. Dadashkarimi. Behavior-Aware COVID-19 Forecasting Using Markov SIR Models on Dynamic Contact Networks: An Observational Modeling Study. medRxiv, 2025.

\bibitem{AlvarezZuzekMuroHavlin2018} L. G. Alvarez-Zuzek, M. A. D. Muro, S. Havlin, et al. Dynamic vaccination in partially overlapped multiplex network. Physical Review E, 2018.

\bibitem{HanYanPei2024} Shixiang Han, Guanghui Yan, Huayan Pei, et al. Dynamical Analysis of an Improved Bidirectional Immunization SIR Model in Complex Network. Entropy, 2024.
\end{thebibliography}
"""
    not_found=return_not_found_references(  references_text=references,
        json_paths=["user-files/question5-12051848/literature_review.json","user-files/question5-12051848/back_grounds.json"],
        threshold_percent=80.0,
        model_name="all-MiniLM-L6-v2"
    )
    print(f"not found reference: {not_found}")
    print("References not found in the JSON library:")
    
    for ref in not_found:
        
        print(f"- {ref}")   